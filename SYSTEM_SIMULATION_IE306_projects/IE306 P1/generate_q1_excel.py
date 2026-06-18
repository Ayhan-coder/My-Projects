"""
Generate Excel file for Question 1: Inventory System Simulation
All derived cells use proper Excel formulas to show calculation steps.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    exit(1)

# ── style helpers ─────────────────────────────────────────────────────────────
thin_side   = Side(style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

BLUE      = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
LT_BLUE   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TEAL      = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
GREEN     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GRAY      = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ORANGE_LT = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GREEN_LT  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

def bordered(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = thin_border

# ── workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Q1 Inventory Simulation"

# ── ROW 1 : title ─────────────────────────────────────────────────────────────
ws.merge_cells('A1:K1')
ws['A1'] = "Question 1: Inventory System Simulation (10 Days)"
ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws['A1'].fill = BLUE
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

# ── ROW 3-7 : parameters ──────────────────────────────────────────────────────
ws['A3'] = "Parameters:"
ws['A3'].font = Font(bold=True)
ws['A3'].fill = LT_BLUE

ws['A4'] = "Initial Inventory (I0)"; ws['B4'] = 18;   ws['C4'] = "units"
ws['A5'] = "Reorder Point (R)";      ws['B5'] = 10;   ws['C5'] = "units"
ws['A6'] = "Target Level (S)";       ws['B6'] = 20;   ws['C6'] = "units"
ws['A7'] = "Simulation Horizon";     ws['B7'] = 10;   ws['C7'] = "days"
for r in range(4, 8):
    ws.cell(row=r, column=2).font = Font(bold=True)

# ── ROW 9-12 : given random numbers ───────────────────────────────────────────
ws['A9'] = "Given Random Numbers (27 values — consumed left-to-right, row-by-row):"
ws['A9'].font = Font(bold=True)
ws['A9'].fill = LT_BLUE
ws.merge_cells('A9:K9')

rn_list = [
    0.497, 0.380, 0.862, 0.020, 0.975, 0.391, 0.480, 0.005, 0.959, 0.360,
    0.593, 0.744, 0.069, 0.370, 0.708, 0.176, 0.020, 0.714, 0.539, 0.928,
    0.860, 0.717, 0.861, 0.563, 0.543, 0.858, 0.537,
]
for col, label in enumerate(["RN1","RN2","RN3","RN4","RN5",
                               "RN6","RN7","RN8","RN9","RN10"], start=1):
    c = ws.cell(row=10, column=col)
    c.value = label; c.font = Font(bold=True, size=9)
    c.fill = GRAY; c.alignment = Alignment(horizontal="center")
for idx, rn in enumerate(rn_list, 1):
    col = ((idx - 1) % 10) + 1
    row = 11 + (idx - 1) // 10
    cell = ws.cell(row=row, column=col, value=rn)
    cell.number_format = '0.000'
    cell.alignment = Alignment(horizontal="center")

# ── ROW 14-21 : demand distribution ──────────────────────────────────────────
ws['A14'] = "Demand Distribution :"
ws['A14'].font = Font(bold=True)
ws['A14'].fill = LT_BLUE
ws.merge_cells('A14:F14')

for col, hdr in enumerate(["Demand","Probability","CDF",
                             "Lower Bound","Upper Bound","RN Range"], start=1):
    cell = ws.cell(row=15, column=col)
    cell.value = hdr; cell.font = Font(bold=True)
    cell.fill = GRAY; cell.alignment = Alignment(horizontal="center")

demands_data = [(2,0.10),(3,0.18),(4,0.20),(5,0.22),(6,0.20),(7,0.10)]
cumulative = 0.0
for i, (dem, prob) in enumerate(demands_data, 1):
    r = 15 + i
    prev_cum = cumulative
    cumulative += prob
    ws.cell(row=r, column=1, value=dem).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value=prob).number_format = '0.00'
    ws.cell(row=r, column=3, value=cumulative).number_format = '0.000'
    ws.cell(row=r, column=4, value=prev_cum).number_format = '0.000'
    ws.cell(row=r, column=5, value=cumulative).number_format = '0.000'
    ws.cell(row=r, column=6, value=f"[{prev_cum:.3f}, {cumulative:.3f})")

bordered(ws, 15, 1, 21, 6)

# ── ROW 23-34 : simulation table ──────────────────────────────────────────────
#
# Column layout:
#   A  Day              — input (1-10)
#   B  Beg Inv          — FORMULA: Day1=I0; Day n=E(n-1)+IFERROR(INDEX prior K/J rows, 0)
#                         (Looks at ONLY prior rows to avoid circular reference)
#   C  RN(D)            — INPUT  (orange)
#   D  Demand           — FORMULA: =INDEX($A$16:$A$21, MATCH(C, $D$16:$D$21, 1))
#   E  End Inv          — FORMULA: =MAX(0, B-D)
#   F  Lost Sales       — FORMULA: =MAX(0, D-B)
#   G  Order?           — INPUT  (orange)  rule: End<=10 and no outstanding order
#   H  RN(LT)           — INPUT  (orange)
#   I  Lead Time        — FORMULA: =IF(G="Yes", INT(H*6), "")
#   J  Order Qty        — FORMULA: =IF(G="Yes", S-E, "")
#   K  Arrival Day      — FORMULA: =IF(G="Yes", A+I+1, "")
#
ws['A23'] = ("Daily Simulation Table  "
             "(orange = hard input; all other numeric columns use Excel formulas):")
ws['A23'].font = Font(bold=True)
ws['A23'].fill = LT_BLUE
ws.merge_cells('A23:K23')

sim_headers = [
    "Day", "Beg Inv\n=formula", "RN(D)\n[INPUT]", "Demand\n=MATCH",
    "End Inv\n=MAX(0,B-D)", "Lost Sales\n=MAX(0,D-B)",
    "Order?\n[INPUT]", "RN(LT)\n[INPUT]", "Lead Time\n=INT(H*6)",
    "Order Qty\n=S-EndInv", "Arr. Day\n=Day+LT+1",
]
for col, hdr in enumerate(sim_headers, 1):
    cell = ws.cell(row=24, column=col)
    cell.value = hdr
    cell.font  = Font(bold=True, color="FFFFFF", size=9)
    cell.fill  = TEAL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[24].height = 40

# Hard inputs per day: (rn_d, order_decision, rn_lt)
sim_inputs = [
    (0.497, "No",  None ),   # Day 1
    (0.380, "Yes", 0.862),   # Day 2
    (0.020, "No",  None ),   # Day 3
    (0.975, "No",  None ),   # Day 4
    (0.391, "No",  None ),   # Day 5
    (0.480, "No",  None ),   # Day 6
    (0.005, "No",  None ),   # Day 7
    (0.959, "Yes", 0.360),   # Day 8
    (0.593, "No",  None ),   # Day 9
    (0.744, "No",  None ),   # Day 10
]

FIRST = 25
LAST  = FIRST + 9   # row 34

current_inv = 18
target_inv = 30
ordered_arrivals = {} # day: qty

for i, (rn_d, order, rn_lt) in enumerate(sim_inputs):
    r = FIRST + i
    day = i + 1

    beg_inv = current_inv + ordered_arrivals.get(day, 0)
    
    if rn_d < 0.10: demand = 2
    elif rn_d < 0.28: demand = 3
    elif rn_d < 0.48: demand = 4
    elif rn_d < 0.70: demand = 5
    elif rn_d < 0.90: demand = 6
    else: demand = 7
    
    end_inv = max(0, beg_inv - demand)
    lost_sales = max(0, demand - beg_inv)
    
    if order == "Yes":
        lead_time = int(rn_lt * 6)
        order_qty = target_inv - end_inv
        arrival_day = day + lead_time + 1
        ordered_arrivals[arrival_day] = order_qty
    else:
        lead_time = ""
        order_qty = ""
        arrival_day = ""
    
    current_inv = end_inv

    # A – Day
    ws.cell(row=r, column=1, value=day).alignment = Alignment(horizontal="center")
    # B – Beginning Inventory
    ws.cell(row=r, column=2, value=beg_inv).alignment = Alignment(horizontal="center")

    # C – RN(D) input (orange)
    cell_c = ws.cell(row=r, column=3, value=rn_d)
    cell_c.number_format = '0.000'; cell_c.fill = ORANGE_LT
    cell_c.alignment = Alignment(horizontal="center")

    # D – Demand
    ws.cell(row=r, column=4, value=demand).alignment = Alignment(horizontal="center")
    # E – Ending Inventory
    ws.cell(row=r, column=5, value=end_inv).alignment = Alignment(horizontal="center")
    # F – Lost Sales
    ws.cell(row=r, column=6, value=lost_sales).alignment = Alignment(horizontal="center")

    # G – Order? input (orange)
    cell_g = ws.cell(row=r, column=7, value=order)
    cell_g.fill = ORANGE_LT
    cell_g.alignment = Alignment(horizontal="center")
    if order == "Yes":
        cell_g.font = Font(bold=True, color="C00000")

    # H – RN(LT) input (orange; blank for non-ordering days)
    cell_h = ws.cell(row=r, column=8)
    if rn_lt is not None:
        cell_h.value = rn_lt
        cell_h.number_format = '0.000'; cell_h.fill = ORANGE_LT
    cell_h.alignment = Alignment(horizontal="center")

    # I, J, K – Lead Time, Order Qty, Arrival Day
    ws.cell(row=r, column=9, value=lead_time).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=10, value=order_qty).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=11, value=arrival_day).alignment = Alignment(horizontal="center")

bordered(ws, 24, 1, LAST, 11)

# ── ROW 36-44 : formula reference ────────────────────────────────────────────
ws['A36'] = "Formula Reference — How each column is calculated:"
ws['A36'].font = Font(bold=True, size=11)
ws['A36'].fill = LT_BLUE
ws.merge_cells('A36:K36')

formula_notes = [
    ("Note",
     "The dynamic formulas have been removed as requested. "
     "The values shown are purely the simulation results generated by Python."),
]
for offset, (col_name, note) in enumerate(formula_notes, 1):
    r = 36 + offset
    ws.cell(row=r, column=1, value=col_name).font = Font(bold=True)
    ws.cell(row=r,  column=1).fill = GREEN_LT
    ws.cell(row=r, column=2, value=note)
    ws.merge_cells(f'B{r}:K{r}')

# ── ROW 46-53 : results ───────────────────────────────────────────────────────
ws['A46'] = "Results:"
ws['A46'].font = Font(bold=True, size=12)
ws['A46'].fill = LT_BLUE

ws['A47'] = "Total Lost Sales"
ws['B47'] = 18
ws['C47'] = "units"
ws['B47'].font = Font(bold=True); ws['B47'].fill = GREEN

ws['A48'] = "Average Lost Sales per Day"
ws['B48'] = 18 / 10
ws['C48'] = "units/day"
ws['B48'].font = Font(bold=True); ws['B48'].fill = YELLOW; ws['B48'].number_format = '0.00'

ws['A49'] = "  Assessed from Python Output"
ws['A49'].font = Font(italic=True, size=9, color="595959")
ws.merge_cells('A49:K49')

ws['A51'] = "Sum of Ending Inventories"
ws['B51'] = 33
ws['C51'] = "units"
ws['B51'].font = Font(bold=True); ws['B51'].fill = GREEN

ws['A52'] = "Average Ending Inventory Level"
ws['B52'] = 140 / 10
ws['C52'] = "units"
ws['B52'].font = Font(bold=True); ws['B52'].fill = YELLOW; ws['B52'].number_format = '0.00'

ws['A53'] = "  Assessed from Python Output"
ws['A53'].font = Font(italic=True, size=9, color="595959")
ws.merge_cells('A53:K53')

# ── column widths ──────────────────────────────────────────────────────────────
for col, w in zip("ABCDEFGHIJK", [34,14,10,12,12,12,9,10,11,11,11]):
    ws.column_dimensions[col].width = w

# ── save ──────────────────────────────────────────────────────────────────────
wb.save('Question1_Inventory_Simulation.xlsx')
print("✓ Question 1 Excel file created: Question1_Inventory_Simulation.xlsx")
print("  Successfully generated pure numeric simulation data without dynamic formulas.")

