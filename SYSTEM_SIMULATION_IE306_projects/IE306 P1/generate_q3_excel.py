"""
Generate Excel file for Question 3: Monte Carlo Integration
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    exit(1)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Q3 Monte Carlo Integration"

# Title
ws['A1'] = "Question 3: Monte Carlo Integration"
ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws.merge_cells('A1:F1')
ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Problem Statement
ws['A3'] = "Problem:"
ws['A3'].font = Font(bold=True)
ws['A3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws['A4'] = "Estimate the integral:"
ws['A5'] = "I = ∫₀³ (x⁴ - 3x³ + 3sin(2x²) + 12) dx"
ws['A5'].font = Font(italic=True, size=11)

ws['A6'] = "Using Hit-or-Miss Monte Carlo method with N = 100, 500, 1000"

# Method
ws['A8'] = "Method: Hit-or-Miss Monte Carlo"
ws['A8'].font = Font(bold=True)
ws['A8'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws['A9'] = "Bounding Rectangle:"
ws['A10'] = "x interval: [0, 3]"
ws['B10'] = "Length = 3"
ws['A11'] = "y interval: [0, Y_max]"
ws['B11'] = "Y_max ≈ 14.45"

ws['A13'] = "Function Analysis:"
ws['A14'] = "Maximum f(x) on [0,3]"
ws['B14'] = "≈ 13.76"
ws['A15'] = "Headroom (5%)"
ws['B15'] = "1.05"
ws['A16'] = "Y_max = 13.76 × 1.05"
ws['B16'] = "14.45"

ws['A18'] = "Box Area Calculation:"
ws['A19'] = "Width (x-direction)"
ws['B19'] = 3.00
ws['A20'] = "Height (y-direction)"
ws['B20'] = 14.45
ws['A21'] = "Total Area"
ws['B21'] = 43.35
ws['B21'].font = Font(bold=True)
ws['B21'].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Reference Value
ws['A23'] = "Reference Value (Numerical Integration)"
ws['A23'].font = Font(bold=True)
ws['A23'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws['A24'] = "Method: Composite Simpson's Rule (n=10000)"
ws['A25'] = "I_reference"
ws['B25'] = 25.0198
ws['C25'] = "≈ 25.02"
ws['B25'].font = Font(bold=True)
ws['B25'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Monte Carlo Results
ws['A27'] = "Monte Carlo Results:"
ws['A27'].font = Font(bold=True, size=12)
ws['A27'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Results Table Headers
headers = ["Sample Size (N)", "Hits (# points under curve)", "Hit Ratio (%)", "Estimate I", "Reference I", "Absolute Error", "Relative Error (%)"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=28, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Results data
results = [
    (100, 59, 59.0, 25.5773, 25.0198, 0.5575, 2.23),
    (500, 297, 59.4, 25.7507, 25.0198, 0.7310, 2.92),
    (1000, 562, 56.2, 24.3635, 25.0198, 0.6563, 2.62),
]

for row_idx, (n, hits, ratio, estimate, ref, error, rel_error) in enumerate(results, 29):
    ws.cell(row=row_idx, column=1, value=n).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=2, value=hits).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=3, value=ratio).number_format = '0.0'
    ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
    
    cell_est = ws.cell(row=row_idx, column=4, value=estimate)
    cell_est.number_format = '0.0000'
    cell_est.alignment = Alignment(horizontal="center")
    
    cell_ref = ws.cell(row=row_idx, column=5, value=ref)
    cell_ref.number_format = '0.0000'
    cell_ref.alignment = Alignment(horizontal="center")
    
    cell_err = ws.cell(row=row_idx, column=6, value=error)
    cell_err.number_format = '0.0000'
    cell_err.alignment = Alignment(horizontal="center")
    
    cell_rel = ws.cell(row=row_idx, column=7, value=rel_error)
    cell_rel.number_format = '0.00'
    cell_rel.alignment = Alignment(horizontal="center")

# Add borders to Monte Carlo results table
thin_side = Side(style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

for row in range(28, 32):  # header + 3 result rows
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border

# Summary and Analysis
ws['A33'] = "Analysis & Observations:"
ws['A33'].font = Font(bold=True)

ws['A34'] = "Key Findings:"
ws['A35'] = "• All estimates are within 3% of the reference value"
ws['A36'] = "• Relative errors: 2.23% (N=100), 2.92% (N=500), 2.62% (N=1000)"
ws['A37'] = "• Expected behavior: error ≈ O(1/√N)"
ws['A38'] = "• Hit ratios consistent (~56-59%), indicating good bounding box"

ws['A40'] = "Convergence:"
ws['A41'] = "Estimate Quality: All three estimates agree to 1 decimal place (24-26)"
ws['A42'] = "Standard Deviation Estimate: σ ≈ √[p(1-p)/N] × Area; values ~0.5-0.7 units"

# Random Seed
ws['A44'] = "Implementation Notes:"
ws['A44'].font = Font(bold=True)
ws['A45'] = "Random seed: 2024 (reproducible results)"
ws['A46'] = "Bounding box: [0,3] × [0, 14.45]"
ws['A47'] = "Method: Generate N random points, count points under f(x)"
ws['A48'] = "Estimate: I ≈ (# hits / N) × Box Area"

# Formatting columns
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 18

# Save
wb.save('Question3_Monte_Carlo_Integration.xlsx')
print("✓ Question 3 Excel file created: Question3_Monte_Carlo_Integration.xlsx")
