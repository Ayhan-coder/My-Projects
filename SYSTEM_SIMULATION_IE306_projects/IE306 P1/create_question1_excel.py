"""
Create Excel simulation for Question 1 - Inventory System
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Q1 Inventory Simulation"

# ── Random numbers ────────────────────────────────
random_numbers = [
    0.497, 0.380, 0.862, 0.020, 0.975, 0.391, 0.480,
    0.005, 0.959, 0.360, 0.593, 0.744, 0.069, 0.370,
    0.708, 0.176, 0.020, 0.714, 0.539, 0.928, 0.860,
    0.717, 0.861, 0.563, 0.543, 0.858, 0.537
]

# ── Parameters ────────────────────────────────────
REORDER_POINT = 10
TARGET_LEVEL = 20
INITIAL_INV = 18
SIM_DAYS = 10

# ── Demand CDF ────────────────────────────────────
# Demand:      2     3     4     5     6     7
# Prob:       0.10  0.18  0.20  0.22  0.20  0.10
# CDF:        0.10  0.28  0.48  0.70  0.90  1.00

# ── Set up title and parameters ────────────────────
title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_font = Font(color="FFFFFF", bold=True, size=12)
header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
header_font = Font(bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws['A1'] = "IE 306 Assignment 1 - Question 1: Inventory System Simulation"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:L1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Parameters section
ws['A3'] = "Parameters"
ws['A3'].font = Font(bold=True, size=11)
ws['A4'] = "Reorder Point:"
ws['B4'] = REORDER_POINT
ws['A5'] = "Target Level:"
ws['B5'] = TARGET_LEVEL
ws['A6'] = "Initial Inventory:"
ws['B6'] = INITIAL_INV
ws['A7'] = "Simulation Days:"
ws['B7'] = SIM_DAYS

# Column headers
headers = ["Day", "Beg Inv", "RN(D)", "Demand", "End Inv", "Lost Sales", 
           "Order?", "RN(LT)", "Lead Time", "Order Qty", "Arr. Day", "Inv Level"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=9, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Simulation logic
rn_idx = 0
inventory = INITIAL_INV
order_pending = False
order_arrival_day = None
order_qty = 0

total_lost_sales = 0
total_ending_inv = 0

row = 10
for day in range(1, SIM_DAYS + 1):
    # Morning: receive order if it arrives
    received = 0
    if order_pending and order_arrival_day == day:
        inventory += order_qty
        order_pending = False
        received = order_qty
        order_qty_val = 0
        order_arrival_day = None

    beginning_inv = inventory

    # During day: generate demand
    rn_demand = random_numbers[rn_idx]
    rn_idx += 1
    
    # Demand from RN
    if rn_demand < 0.10:
        demand = 2
    elif rn_demand < 0.28:
        demand = 3
    elif rn_demand < 0.48:
        demand = 4
    elif rn_demand < 0.70:
        demand = 5
    elif rn_demand < 0.90:
        demand = 6
    else:
        demand = 7

    # Satisfy demand
    if inventory >= demand:
        lost = 0
        inventory -= demand
    else:
        lost = demand - inventory
        inventory = 0

    ending_inv = inventory
    total_lost_sales += lost
    total_ending_inv += ending_inv

    # End of day: check reorder policy
    order_placed = False
    rn_lt = ""
    lt = ""
    order_qty_str = ""
    arrive_day = ""

    if ending_inv <= REORDER_POINT and not order_pending:
        order_placed = True
        order_qty = TARGET_LEVEL - ending_inv
        rn_lt = random_numbers[rn_idx]
        rn_idx += 1
        
        # Lead time from RN (discrete uniform 0-5)
        if rn_lt < 1/6:
            lt = 0
        elif rn_lt < 2/6:
            lt = 1
        elif rn_lt < 3/6:
            lt = 2
        elif rn_lt < 4/6:
            lt = 3
        elif rn_lt < 5/6:
            lt = 4
        else:
            lt = 5

        order_pending = True
        order_arrival_day = day + lt + 1
        order_qty_str = order_qty
        arrive_day = order_arrival_day
    else:
        rn_lt = ""
        lt = ""
        order_qty_str = ""
        arrive_day = ""

    # Write row
    ws.cell(row=row, column=1).value = day
    ws.cell(row=row, column=2).value = beginning_inv + received if received > 0 else beginning_inv
    ws.cell(row=row, column=3).value = round(rn_demand, 3) if rn_demand else ""
    ws.cell(row=row, column=4).value = demand if rn_demand else ""
    ws.cell(row=row, column=5).value = ending_inv
    ws.cell(row=row, column=6).value = lost
    ws.cell(row=row, column=7).value = "Yes" if order_placed else "No"
    ws.cell(row=row, column=8).value = round(rn_lt, 3) if rn_lt != "" else ""
    ws.cell(row=row, column=9).value = lt if lt != "" else ""
    ws.cell(row=row, column=10).value = order_qty_str if order_qty_str != "" else ""
    ws.cell(row=row, column=11).value = arrive_day if arrive_day != "" else ""
    ws.cell(row=row, column=12).value = ending_inv

    # Format cells
    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row += 1

# Summary statistics
summary_row = row + 2
ws.cell(row=summary_row, column=1).value = "RESULTS"
ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)

summary_row += 1
ws.cell(row=summary_row, column=1).value = "Total Lost Sales:"
ws.cell(row=summary_row, column=2).value = total_lost_sales

summary_row += 1
ws.cell(row=summary_row, column=1).value = "Average Lost Sales per Day:"
ws.cell(row=summary_row, column=2).value = round(total_lost_sales / SIM_DAYS, 2)

summary_row += 1
ws.cell(row=summary_row, column=1).value = "Sum of Ending Inventories:"
ws.cell(row=summary_row, column=2).value = total_ending_inv

summary_row += 1
ws.cell(row=summary_row, column=1).value = "Average Inventory Level:"
ws.cell(row=summary_row, column=2).value = round(total_ending_inv / SIM_DAYS, 2)

# Format summary section
for i in range(4):
    ws.cell(row=row + 2 + i, column=1).font = Font(bold=True)

# Column widths
ws.column_dimensions['A'].width = 15
for col in range(2, 13):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Save
wb.save('Question1_Inventory_Simulation.xlsx')
print("✓ Question 1 Excel file created: Question1_Inventory_Simulation.xlsx")
