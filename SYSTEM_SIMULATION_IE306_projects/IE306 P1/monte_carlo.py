import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def integrand(x):
    """The function to integrate: x^4 - 3x^3 + 3*sin(2x^2) + 12"""
    return x**4 - 3*x**3 + 3*np.sin(2*x**2) + 12

def hit_or_miss_mc(N, x_min=0, x_max=3, y_min=0, y_max=14.45):
    """
    Hit-or-Miss Monte Carlo integration
    
    Parameters:
    N: number of random points
    x_min, x_max: integration bounds
    y_min, y_max: bounding rectangle bounds
    
    Returns:
    estimate: estimated integral value
    hits: number of points under the curve
    """
    # Generate N random points in the rectangle
    x_points = np.random.uniform(x_min, x_max, N)
    y_points = np.random.uniform(y_min, y_max, N)
    
    # Count hits (points under the curve)
    hits = np.sum(y_points <= integrand(x_points))
    
    # Estimate integral
    rectangle_area = (x_max - x_min) * (y_max - y_min)
    estimate = rectangle_area * (hits / N)
    
    return estimate, hits

def run_monte_carlo():
    """Run Monte Carlo integration for different sample sizes"""
    print("Monte Carlo Integration")
    print("=" * 60)
    
    # Reference value using composite Simpson's rule
    def simpsons_quad(func, a, b, n=10000):
        if n % 2 == 1:
            n += 1
        h = (b - a) / n
        xs = np.linspace(a, b, n + 1)
        ys = func(xs)
        return (h / 3) * (ys[0] + ys[-1] + 4 * np.sum(ys[1::2]) + 2 * np.sum(ys[2:-1:2]))

    reference = simpsons_quad(integrand, 0, 3)
    print(f"Reference value (numerical integration): {reference:.6f}")
    
    # Find maximum value for bounding rectangle
    x_vals = np.linspace(0, 3, 1000)
    y_vals = integrand(x_vals)
    y_max = np.max(y_vals)
    print(f"Maximum function value on [0,3]: {y_max:.6f}")
    
    # Run Monte Carlo for different sample sizes
    sample_sizes = [100, 500, 1000]
    results = []
    
    print(f"\n{'N':<6} {'Hits':<6} {'Estimate':<10} {'Abs Error':<10} {'Rel Error (%)':<12}")
    print("-" * 60)
    
    np.random.seed(42)  # For reproducible results
    
    for N in sample_sizes:
        estimate, hits = hit_or_miss_mc(N)
        abs_error = abs(estimate - reference)
        rel_error = (abs_error / reference) * 100
        
        results.append({
            'N': N,
            'hits': hits,
            'estimate': estimate,
            'abs_error': abs_error,
            'rel_error': rel_error
        })
        
        print(f"{N:<6} {hits:<6} {estimate:<10.6f} {abs_error:<10.6f} {rel_error:<12.2f}")
    
    return results, reference

if __name__ == "__main__":
    results, reference = run_monte_carlo()
    
    # Generate Excel file if openpyxl is available
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "Monte Carlo Integration"
        
        # Title
        ws['A1'] = "Monte Carlo Integration Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Problem
        ws['A3'] = "Integral: I = ∫₀³ (x⁴ - 3x³ + 3sin(2x²) + 12) dx"
        ws['A3'].font = Font(italic=True)
        
        # Reference
        ws['A5'] = "Reference Value (Simpson's Rule):"
        ws['B5'] = reference
        ws['B5'].number_format = '0.0000'
        
        # Bounding Box
        ws['A7'] = "Bounding Box: x ∈ [0,3], y ∈ [0, 14.45]"
        ws['A8'] = "Box Area: 3 × 14.45 = 43.35"
        
        # Results Table
        ws['A10'] = "Results:"
        ws['A10'].font = Font(bold=True, size=12)
        
        headers = ["Sample Size (N)", "Hits", "Estimate", "Reference", "Abs Error", "Rel Error (%)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, result in enumerate(results, 12):
            ws.cell(row=row_idx, column=1, value=result['N']).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=result['hits']).alignment = Alignment(horizontal="center")
            
            cell = ws.cell(row=row_idx, column=3, value=result['estimate'])
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            
            cell = ws.cell(row=row_idx, column=4, value=reference)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            
            cell = ws.cell(row=row_idx, column=5, value=result['abs_error'])
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            
            cell = ws.cell(row=row_idx, column=6, value=result['rel_error'])
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center")
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        
        wb.save('Monte_Carlo_Results.xlsx')
        print("\n✓ Excel file generated: Monte_Carlo_Results.xlsx")
    else:
        print("\nNote: openpyxl not installed. Install with: pip install openpyxl")
